from rest_framework import viewsets, filters
from rest_framework.exceptions import ValidationError
from users.permissions import AdminOrReadOnly, RolePermission
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db import transaction

from schools.models import School
from .models import ClassRoom, Section, Subject, StudentProfile, TeacherAssignment
from .serializers import (
    SchoolSerializer, ClassRoomSerializer, SectionSerializer,
    SubjectSerializer, StudentProfileSerializer, TeacherAssignmentSerializer
)

from results.models import Examination, StudentOverallResult
from django.db.models import Q, Count


class SchoolListAPI(APIView):
    def get(self, request):
        schools = School.objects.all()
        serializer = SchoolSerializer(schools, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class ClassRoomViewSet(viewsets.ModelViewSet):
    queryset = ClassRoom.objects.select_related('school').prefetch_related('sections', 'students').all()
    serializer_class = ClassRoomSerializer
    permission_classes = [AdminOrReadOnly]
    filter_backends = [filters.SearchFilter]
    search_fields = ['name']
    
    def get_queryset(self):
        queryset = ClassRoom.objects.select_related('school').prefetch_related('sections', 'students').all()
        school_id = self.request.query_params.get('school')
        if school_id:
            queryset = queryset.filter(school_id=school_id)
        return queryset
    
    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Get class summary with student counts"""
        school_id = request.query_params.get('school')
        if not school_id:
            return Response({"detail": "school parameter required"}, status=status.HTTP_400_BAD_REQUEST)
        
        classrooms = ClassRoom.objects.filter(school_id=school_id).prefetch_related('students')
        data = []
        for classroom in classrooms:
            data.append({
                'id': classroom.id,
                'name': classroom.name,
                'description': classroom.description,
                'student_count': classroom.students.count(),
                'subject_count': Subject.objects.filter(school_id=school_id).count()  # subjects are school-wide
            })
        return Response(data)
    
    @action(detail=True, methods=['get'])
    def students(self, request, pk=None):
        """Get all students in a specific class"""
        classroom = self.get_object()
        students = StudentProfile.objects.filter(classroom=classroom).select_related('user', 'guardian', 'section')
        serializer = StudentProfileSerializer(students, many=True, context={'request': request})
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def subjects(self, request, pk=None):
        classroom = self.get_object()
        from django.db.models import Q
        subjects = Subject.objects.filter(school=classroom.school).filter(
            Q(classrooms=classroom) | Q(assignments__classroom=classroom)
        ).distinct()
        name = classroom.name or ''
        import re
        g = None
        m = re.search(r'(\d+)', name)
        if m:
            try:
                g = int(m.group(1))
            except Exception:
                g = None
        if g is None:
            bn_map = {'প্রথম':1,'দ্বিতীয়':2,'তৃতীয়':3,'চতুর্থ':4,'পঞ্চম':5,'ষষ্ঠ':6,'সপ্তম':7,'অষ্টম':8,'নবম':9,'দশম':10,'একাদশ':11,'দ্বাদশ':12}
            for k,v in bn_map.items():
                if k in name:
                    g = v
                    break
        if g and g <= 8:
            subjects = subjects.exclude(name__iregex=r'(physics|পদার্থ|chemistry|রসায়ন)')
        
        result = []
        for subject in subjects:
            assignments = TeacherAssignment.objects.filter(
                subject=subject,
                classroom=classroom
            ).select_related('teacher')
            
            teachers = []
            for assignment in assignments:
                teacher = assignment.teacher
                teachers.append({
                    'id': teacher.id,
                    'name': f"{teacher.first_name} {teacher.last_name}".strip() or teacher.username,
                    'username': teacher.username
                })
            
            result.append({
                'id': subject.id,
                'name': subject.name,
                'code': subject.code,
                'teachers': teachers,
                'notifications': 0  # Placeholder for future notification count
            })
        
        return Response(result)


class SectionViewSet(viewsets.ModelViewSet):
    queryset = Section.objects.select_related('classroom__school').all()
    serializer_class = SectionSerializer
    permission_classes = [RolePermission]
    filter_backends = []
    
    def get_queryset(self):
        queryset = Section.objects.select_related('classroom__school').all()
        classroom_id = self.request.query_params.get('classroom')
        school_id = self.request.query_params.get('school')
        if classroom_id:
            queryset = queryset.filter(classroom_id=classroom_id)
        if school_id:
            queryset = queryset.filter(classroom__school_id=school_id)
        return queryset
    
    def create(self, request, *args, **kwargs):
        classroom_id = request.data.get('classroom_id')
        name = (request.data.get('name') or '').strip()
        if not classroom_id or not name:
            return super().create(request, *args, **kwargs)
        try:
            count = Section.objects.filter(classroom_id=classroom_id).count()
            # Optional: keep a soft cap to avoid explosion, but do not hard-fail for groups like বিজ্ঞান/মানবিক/ব্যবসায়
            if count >= 12:
                return Response({"detail": "Too many sections for this class"}, status=status.HTTP_400_BAD_REQUEST)
        except Exception:
            pass
        return super().create(request, *args, **kwargs)


class SubjectViewSet(viewsets.ModelViewSet):
    queryset = Subject.objects.select_related('school').all()
    serializer_class = SubjectSerializer
    permission_classes = [AdminOrReadOnly]
    filter_backends = [filters.SearchFilter]
    search_fields = ['name','code']
    
    def get_queryset(self):
        queryset = Subject.objects.select_related('school').prefetch_related('classrooms').all()
        school_id = self.request.query_params.get('school')
        classroom_id = self.request.query_params.get('classroom')
        if school_id:
            queryset = queryset.filter(school_id=school_id)
        # We don't strictly filter by classroom here if we want to show all school subjects
        # but if the frontend explicitly asks for classroom subjects, we can filter.
        # However, for the "SubjectsPage" we usually fetch all and filter in frontend OR
        # fetch per classroom. Let's support filtering.
        if classroom_id:
            queryset = queryset.filter(classrooms__id=classroom_id)
        return queryset
    
    def create(self, request, *args, **kwargs):
        school_id = request.data.get('school_id')
        name = request.data.get('name')
        if school_id and name:
            existing_subject = Subject.objects.filter(school_id=school_id, name=name).first()
            if existing_subject:
                classrooms = request.data.get('classrooms', [])
                classroom_id = request.data.get('classroom_id')
                sections = request.data.get('sections', [])
                ids_to_add = set()
                if isinstance(classrooms, list):
                    try:
                        ids_to_add.update([int(c) for c in classrooms])
                    except (ValueError, TypeError):
                        pass
                if classroom_id:
                    try:
                        ids_to_add.add(int(classroom_id))
                    except (ValueError, TypeError):
                        pass
                if ids_to_add:
                    existing_subject.classrooms.add(*ids_to_add)
                section_ids = set()
                if isinstance(sections, list):
                    try:
                        section_ids.update([int(s) for s in sections])
                    except (ValueError, TypeError):
                        pass
                if section_ids:
                    existing_subject.sections.add(*section_ids)
                code = request.data.get('code')
                if code and code != existing_subject.code:
                    existing_subject.code = code
                    existing_subject.save()
                serializer = self.get_serializer(existing_subject)
                return Response(serializer.data, status=status.HTTP_200_OK)
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        classroom_id = self.request.data.get('classroom_id')
        instance = serializer.save()
        if classroom_id:
            try:
                classroom = ClassRoom.objects.get(id=classroom_id)
                instance.classrooms.add(classroom)
            except ClassRoom.DoesNotExist:
                pass

    @action(detail=True, methods=['post'])
    def assign_class(self, request, pk=None):
        """Assign this subject to a classroom"""
        subject = self.get_object()
        classroom_id = request.data.get('classroom_id')
        if not classroom_id:
            return Response({"detail": "classroom_id required"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            classroom = ClassRoom.objects.get(id=classroom_id)
            subject.classrooms.add(classroom)
            return Response({"status": "assigned"})
        except ClassRoom.DoesNotExist:
            return Response({"detail": "Classroom not found"}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['get'])
    def detail(self, request, pk=None):
        """Get detailed subject information including assignments, results, and attendance"""
        subject = self.get_object()
        
        # Get teacher assignments
        assignments = TeacherAssignment.objects.filter(subject=subject).select_related(
            'teacher', 'classroom', 'section'
        )
        
        assignments_data = []
        for assignment in assignments:
            teacher = assignment.teacher
            assignments_data.append({
                'id': assignment.id,
                'teacher': {
                    'id': teacher.id,
                    'name': f"{teacher.first_name} {teacher.last_name}".strip() or teacher.username,
                    'username': teacher.username
                },
                'classroom': {
                    'id': assignment.classroom.id,
                    'name': assignment.classroom.name
                },
                'section': {
                    'id': assignment.section.id,
                    'name': assignment.section.name
                } if assignment.section else None
            })
        
        # Get recent results for this subject
        from results.models import Result
        recent_results = Result.objects.filter(subject=subject).select_related(
            'examination', 'student__user'
        ).order_by('-examination__exam_date')[:20]
        
        results_data = []
        for result in recent_results:
            results_data.append({
                'examination': result.examination.name,
                'student': f"{result.student.user.first_name} {result.student.user.last_name}".strip() or result.student.user.username,
                'total_obtained': float(result.total_obtained),
                'grade': result.grade
            })
        
        return Response({
            'id': subject.id,
            'name': subject.name,
            'code': subject.code,
            'assignments': assignments_data,
            'recent_results': results_data,
            'total_assignments': len(assignments_data),
            'notifications': 0  # Placeholder
        })


class StudentProfileViewSet(viewsets.ModelViewSet):
    queryset = StudentProfile.objects.select_related('user', 'school', 'classroom', 'section', 'guardian').all()
    serializer_class = StudentProfileSerializer
    permission_classes = [RolePermission]
    filter_backends = [filters.SearchFilter]
    search_fields = [
        'user__username',
        'user__first_name',
        'user__last_name',
        'roll_number',
        'guardian_name',
        'guardian__first_name',
        'guardian__last_name'
    ]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except ValidationError as e:
            return Response(e.detail, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            msg = str(e)
            return Response({'detail': msg or 'Unable to create student'}, status=status.HTTP_400_BAD_REQUEST)
    
    def get_queryset(self):
        queryset = StudentProfile.objects.select_related('user', 'school', 'classroom', 'section', 'guardian').all()
        school_id = self.request.query_params.get('school')
        classroom_id = self.request.query_params.get('classroom')
        section_id = self.request.query_params.get('section')
        if school_id:
            queryset = queryset.filter(school_id=school_id)
        if classroom_id:
            queryset = queryset.filter(classroom_id=classroom_id)
        if section_id:
            queryset = queryset.filter(section_id=section_id)
            
        # Sort by roll number (numeric), with empty/nulls at the end
        # Moved sorting to frontend to avoid DB-specific regex issues (e.g. SQLite vs Postgres)
        queryset = queryset.order_by('roll_number')
        
        return queryset
    
    def _validate_section(self, data):
        """Ensure section belongs to classroom"""
        section_id = data.get('section_id')
        classroom_id = data.get('classroom_id')
        if section_id:
            try:
                section = Section.objects.select_related('classroom').get(id=section_id)
            except Section.DoesNotExist:
                raise ValidationError({"section_id": "Invalid section"})
            if classroom_id and int(section.classroom_id) != int(classroom_id):
                raise ValidationError({"section_id": "Section must belong to the selected classroom"})
    
    def perform_create(self, serializer):
        self._validate_section(self.request.data)
        serializer.save()
    
    def perform_update(self, serializer):
        self._validate_section(self.request.data)
        serializer.save()
    
    @action(detail=True, methods=['get'])
    def detail(self, request, pk=None):
        """Get detailed student information including results and attendance"""
        student = self.get_object()
        
        # Basic student info
        from users.serializers import UserSerializer
        user_data = UserSerializer(student.user, context={'request': request}).data
        
        # Get recent results
        from results.models import Result
        recent_results = Result.objects.filter(student=student).select_related(
            'examination', 'subject'
        ).order_by('-examination__exam_date')[:10]
        
        results_data = []
        for result in recent_results:
            results_data.append({
                'examination': result.examination.name,
                'subject': result.subject.name,
                'total_obtained': float(result.total_obtained),
                'grade': result.grade,
                'gpa': float(result.gpa),
                'is_passed': result.is_passed
            })
        
        # Get attendance summary (placeholder - implement when attendance app is ready)
        attendance_summary = {
            'total_days': 0,
            'present_days': 0,
            'absent_days': 0,
            'percentage': 0
        }
        
        return Response({
            'id': student.id,
            'user': user_data,
            'classroom': {'id': student.classroom.id, 'name': student.classroom.name} if student.classroom else None,
            'section': {'id': student.section.id, 'name': student.section.name} if student.section else None,
            'roll_number': student.roll_number,
            'guardian_name': student.guardian_name,
            'guardian': UserSerializer(student.guardian, context={'request': request}).data if student.guardian else None,
            'recent_results': results_data,
            'attendance': attendance_summary
        })
    
    @action(detail=True, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def upload_photo(self, request, pk=None):
        """Upload photo for a student"""
        student = self.get_object()
        
        if 'photo' in request.FILES:
            student.user.photo = request.FILES['photo']
            student.user.save()
            
            from users.serializers import UserSerializer
            return Response({
                "message": "Photo uploaded successfully",
                "user": UserSerializer(student.user, context={'request': request}).data
            })
        
        return Response({"error": "No photo provided"}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Export students to CSV"""
        import csv
        from django.http import HttpResponse
        
        school_id = request.query_params.get('school')
        classroom_id = request.query_params.get('classroom')
        
        qs = self.filter_queryset(self.get_queryset())
        
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="students_export.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Serial', 'Student Name', 'Username', 'Class', 'Section', 'Roll Number', 'Parent Name', 'Parent Username'])
        
        for idx, sp in enumerate(qs, start=1):
            user = sp.user
            student_name = f"{user.first_name} {user.last_name}".strip() or user.username
            classroom = sp.classroom.name if sp.classroom else ''
            section = sp.section.name if sp.section else ''
            parent_name = sp.guardian_name or ''
            parent_username = sp.guardian.username if sp.guardian else ''
            
            writer.writerow([idx, student_name, user.username, classroom, section, sp.roll_number or '', parent_name, parent_username])
        
        return response

    @action(detail=False, methods=['post'], permission_classes=[AdminOrReadOnly])
    def purge_no_roll(self, request):
        """
        Delete all students who have no roll number across all schools,
        including their linked user and all related records (results, attendance, fees).
        Optional: pass ?school=<id> to limit to a single school.
        """
        school_id = request.data.get('school') or request.query_params.get('school')
        base_qs = StudentProfile.objects.all()
        if school_id:
            base_qs = base_qs.filter(school_id=school_id)
        targets = base_qs.filter(Q(roll_number__isnull=True) | Q(roll_number__exact=''))
        total = targets.count()
        if total == 0:
            return Response({"status": "ok", "deleted": 0, "by_school": []}, status=status.HTTP_200_OK)
        stats = list(targets.values('school_id').annotate(deleted=Count('id')).order_by('school_id'))
        with transaction.atomic():
            targets.delete()
        return Response({"status": "ok", "deleted": total, "by_school": stats}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def promote(self, request):
        """
        Promote students to next class based on FINAL (annual) examination result.
        Rules:
        - Class 8 -> Class 9: put ALL promoted students into a single section ('ক')
        - Class 10 -> 'S.S.C. পরীক্ষার্থী' class
        - Others: promote to next class, preserving section name (ক/খ/গ) when available
        """
        school_id = request.data.get('school') or request.query_params.get('school')
        if not school_id:
            return Response({"detail": "school parameter is required"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            school = School.objects.get(pk=school_id)
        except School.DoesNotExist:
            return Response({"detail": "Invalid school id"}, status=status.HTTP_400_BAD_REQUEST)

        # Helpers
        import re
        bn_map = {'প্রথম':1,'দ্বিতীয়':2,'তৃতীয়':3,'চতুর্থ':4,'পঞ্চম':5,'ষষ্ঠ':6,'সপ্তম':7,'অষ্টম':8,'নবম':9,'দশম':10,'একাদশ':11,'দ্বাদশ':12}
        def parse_grade(class_name):
            if not class_name:
                return None
            m = re.search(r'(\d+)', class_name)
            if m:
                try:
                    return int(m.group(1))
                except Exception:
                    pass
            for k, v in bn_map.items():
                if k in class_name:
                    return v
            return None

        def get_or_create_section(classroom, name):
            try:
                sec = Section.objects.get(classroom=classroom, name=name)
            except Section.DoesNotExist:
                sec = Section.objects.create(classroom=classroom, name=name)
            return sec

        # Build class lookup by grade and also SSC special class
        classes = list(ClassRoom.objects.filter(school=school))
        by_grade = {}
        ssc_class = None
        for c in classes:
            g = parse_grade(c.name or '')
            if g:
                by_grade.setdefault(g, []).append(c)
            name_l = (c.name or '').lower()
            if ('s.s.c' in name_l) or ('ssc' in name_l) or ('এস.এস.সি' in c.name) or ('এসএসসি' in c.name) or ('পরীক্ষার্থী' in c.name):
                ssc_class = c
        if ssc_class is None:
            ssc_class = ClassRoom.objects.create(school=school, name='S.S.C. পরীক্ষার্থী')
            classes.append(ssc_class)

        # For determinism, when multiple classes match a grade, pick the first by name
        for g in list(by_grade.keys()):
            by_grade[g].sort(key=lambda x: x.name or '')

        promoted, skipped_no_exam, skipped_failed = 0, 0, 0
        with transaction.atomic():
            students = StudentProfile.objects.filter(school=school).select_related('classroom','section','user')
            for sp in students:
                cls = sp.classroom
                if not cls:
                    continue
                grade = parse_grade(cls.name or '')
                if not grade:
                    continue
                # Find latest 'annual' exam for the student's class (match classroom)
                exam_qs = Examination.objects.filter(
                    school=school,
                    classroom=cls
                ).filter(
                    Q(exam_type='annual') |
                    Q(name__icontains='final') |
                    Q(name__icontains='বার্ষিক') |
                    Q(name__icontains='ফাইনাল')
                ).order_by('-exam_date', '-id')
                exam = exam_qs.first()
                if not exam:
                    skipped_no_exam += 1
                    continue
                overall = StudentOverallResult.objects.filter(examination=exam, student=sp).first()
                if not overall or not overall.is_passed:
                    skipped_failed += 1
                    continue

                # Determine target classroom and section
                target_class = None
                target_section = None
                if grade == 10:
                    target_class = ssc_class
                    # Preserve section if exists; else create 'ক'
                    sec_name = (sp.section.name if sp.section else 'ক')
                    target_section = get_or_create_section(target_class, sec_name)
                else:
                    target_grade = grade + 1
                    options = by_grade.get(target_grade, [])
                    if not options:
                        # If next grade class doesn't exist, create a new one
                        # Name convention: use Bengali names when possible
                        rev_bn = {v: k for k, v in bn_map.items()}
                        name = rev_bn.get(target_grade, f"Class {target_grade}")
                        target_class = ClassRoom.objects.create(school=school, name=name)
                        by_grade.setdefault(target_grade, []).append(target_class)
                    else:
                        target_class = options[0]
                    # Preserve section name if possible for ALL grades
                    # Create target section with same name when needed
                    sec_name = sp.section.name if sp.section else 'ক'
                    target_section = get_or_create_section(target_class, sec_name)

                # Apply promotion
                sp.classroom = target_class
                sp.section = target_section
                sp.save(update_fields=['classroom','section'])
                promoted += 1

        return Response({
            "status": "ok",
            "school": school.id,
            "promoted": promoted,
            "skipped_no_exam": skipped_no_exam,
            "skipped_failed": skipped_failed
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'])
    def promote_class(self, request):
        """
        Promote PASSING students from one class to another based on Annual/Final exam.
        Params:
        - school (required)
        - from_class_id (required)
        - to_class_id (required)
        - section_mode: 'preserve' | 'single' (default: 'preserve')
        - single_section_name: when section_mode == 'single', default 'ক'
        - Optional: exam_id (override detection)
        Behavior (Result-based):
        - Detect latest Annual/Final exam for from_class (or use exam_id)
        - Move only students who PASSED (overall) to to_class_id
        - If section_mode='preserve', creates matching section names in target as needed
        - If section_mode='single', puts everyone into the given single section
        """
        school_id = request.data.get('school') or request.query_params.get('school')
        from_class_id = request.data.get('from_class_id')
        to_class_id = request.data.get('to_class_id')
        section_mode = (request.data.get('section_mode') or 'preserve').strip()
        single_section_name = (request.data.get('single_section_name') or 'ক').strip() or 'ক'
        explicit_exam_id = request.data.get('exam_id')
        if not school_id or not from_class_id or not to_class_id:
            return Response({"detail": "school, from_class_id and to_class_id are required"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            school = School.objects.get(pk=school_id)
        except School.DoesNotExist:
            return Response({"detail": "Invalid school id"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            from_class = ClassRoom.objects.get(pk=from_class_id, school=school)
            to_class = ClassRoom.objects.get(pk=to_class_id, school=school)
        except ClassRoom.DoesNotExist:
            return Response({"detail": "Invalid from_class_id or to_class_id for this school"}, status=status.HTTP_400_BAD_REQUEST)

        # Determine the exam: explicit, or by exam_type/year, or auto-detected "final/annual"
        exam = None
        if explicit_exam_id:
            exam = Examination.objects.filter(pk=explicit_exam_id, school=school, classroom=from_class).first()
        if not exam:
            req_exam_type = (request.data.get('exam_type') or request.query_params.get('exam_type') or '').strip().lower()
            req_year = request.data.get('year') or request.query_params.get('year')
            qs = Examination.objects.filter(school=school, classroom=from_class)
            # Optional: filter by year if provided
            try:
                if req_year:
                    qs = qs.filter(exam_date__year=int(req_year))
            except Exception:
                pass
            # If a specific exam_type requested, try that first
            if req_exam_type:
                # Attempt direct match
                exam = qs.filter(exam_type=req_exam_type).order_by('-exam_date', '-id').first()
                # Fallbacks for common aliases
                if not exam and req_exam_type in ('final', 'annual'):
                    exam = qs.filter(
                        Q(exam_type='annual') |
                        Q(name__icontains='final') |
                        Q(name__icontains='বার্ষিক') |
                        Q(name__icontains='ফাইনাল')
                    ).order_by('-exam_date', '-id').first()
            # If still not found, auto-detect annual/final by name/type
            if not exam:
                exam = qs.filter(
                    Q(exam_type='annual') |
                    Q(name__icontains='final') |
                    Q(name__icontains='বার্ষিক') |
                    Q(name__icontains='ফাইনাল')
                ).order_by('-exam_date', '-id').first()

        # If no exam found, create snapshot with 'retained' and do not move anyone
        if not exam:
            from django.utils import timezone
            total_candidates = StudentProfile.objects.filter(school=school, classroom=from_class).count()
            year = str(timezone.now().year)
            students = StudentProfile.objects.filter(school=school, classroom=from_class).select_related('section')
            for sp in students:
                try:
                    from academics.models import StudentYearRecord
                    StudentYearRecord.objects.update_or_create(
                        student=sp,
                        academic_year=year,
                        defaults={
                            'school_id': int(school.id),
                            'classroom': sp.classroom,
                            'section': sp.section,
                            'roll_number': sp.roll_number,
                            'status': 'retained',
                            'promoted_to_classroom': None,
                            'examination_id': None,
                            'result_cgpa': None,
                            'result_grade': None,
                            'percentage': None,
                            'rank': None,
                            'promoted_on': None,
                            'meta': {'note': 'No annual exam found'}
                        }
                    )
                except Exception:
                    pass
            return Response({
                "status": "ok",
                "school": school.id,
                "from_class": from_class.id,
                "to_class": to_class.id,
                "exam_found": False,
                "moved": 0,
                "total_candidates": total_candidates,
                "skipped_no_exam": total_candidates
            }, status=status.HTTP_200_OK)

        # Passed students for this exam
        passer_ids = set(StudentOverallResult.objects.filter(
            examination=exam,
            is_passed=True
        ).values_list('student_id', flat=True))

        def get_or_create_section(classroom, name):
            try:
                return Section.objects.get(classroom=classroom, name=name)
            except Section.DoesNotExist:
                return Section.objects.create(classroom=classroom, name=name)

        from django.utils import timezone
        total_candidates = StudentProfile.objects.filter(school=school, classroom=from_class).count()
        moved = 0
        with transaction.atomic():
            # Snapshot all candidates with result details
            all_candidates = StudentProfile.objects.filter(
                school=school,
                classroom=from_class
            ).select_related('section')
            year = str(exam.exam_date.year) if exam.exam_date else str(timezone.now().year)
            for sp in all_candidates:
                overall = StudentOverallResult.objects.filter(examination=exam, student=sp).first()
                try:
                    from academics.models import StudentYearRecord
                    StudentYearRecord.objects.update_or_create(
                        student=sp,
                        academic_year=year,
                        defaults={
                            'school_id': int(school.id),
                            'classroom': sp.classroom,
                            'section': sp.section,
                            'roll_number': sp.roll_number,
                            'status': 'promoted' if (overall and overall.is_passed) else ('not_passed' if overall else 'retained'),
                            'promoted_to_classroom': to_class if (overall and overall.is_passed) else None,
                            'examination_id': exam.id,
                            'result_cgpa': getattr(overall, 'cgpa', None),
                            'result_grade': getattr(overall, 'grade', None),
                            'percentage': getattr(overall, 'percentage', None),
                            'rank': getattr(overall, 'rank', None),
                            'promoted_on': timezone.now() if (overall and overall.is_passed) else None,
                            'meta': {'exam_name': exam.name, 'exam_type': exam.exam_type}
                        }
                    )
                except Exception:
                    pass
            # Move only passers
            students = all_candidates.filter(id__in=list(passer_ids))
            if section_mode == 'single':
                target_section = get_or_create_section(to_class, single_section_name)
                moved = students.update(classroom=to_class, section=target_section)
            else:
                # preserve sections
                for sp in students:
                    sec_name = sp.section.name if sp.section else 'ক'
                    tgt_sec = get_or_create_section(to_class, sec_name)
                    sp.classroom = to_class
                    sp.section = tgt_sec
                    sp.save(update_fields=['classroom', 'section'])
                    moved += 1

        skipped_not_passed = max(total_candidates - moved, 0)
        return Response({
            "status": "ok",
            "school": school.id,
            "from_class": from_class.id,
            "to_class": to_class.id,
            "exam_found": True,
            "exam_id": exam.id,
            "exam_name": exam.name,
            "moved": moved,
            "total_candidates": total_candidates,
            "skipped_not_passed": skipped_not_passed
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'])
    def year_report(self, request):
        """
        Revert promotion using StudentYearRecord snapshots.
        Params:
        - school (required)
        - year (required): academic year string or int
        - target_class_id (optional): only revert those currently promoted_to_classroom == target
        - student_ids (optional): JSON list of student ids to restrict
        Behavior:
        - For matching StudentYearRecord entries with status='promoted', move student back to the snapshot 'classroom' and 'section'
        """
        school_id = request.data.get('school') or request.query_params.get('school')
        year = request.data.get('year') or request.query_params.get('year')
        target_class_id = request.data.get('target_class_id') or request.query_params.get('target_class_id')
        student_ids = request.data.get('student_ids') or request.query_params.get('student_ids')
        if not school_id or not year:
            return Response({"detail": "school and year are required"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            school = School.objects.get(pk=school_id)
        except School.DoesNotExist:
            return Response({"detail": "Invalid school id"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            from academics.models import StudentYearRecord
            qs = StudentYearRecord.objects.select_related('student', 'classroom', 'section', 'promoted_to_classroom') \
                .filter(school_id=school_id, academic_year=str(year), status='promoted')
            if target_class_id:
                qs = qs.filter(promoted_to_classroom_id=target_class_id)
            # Restrict to provided student ids if any
            ids_set = None
            if student_ids:
                try:
                    if isinstance(student_ids, str):
                        import json
                        ids_set = set(json.loads(student_ids))
                    elif isinstance(student_ids, list):
                        ids_set = set(student_ids)
                except Exception:
                    pass
                if ids_set:
                    qs = qs.filter(student_id__in=list(ids_set))
            records = list(qs)
            reverted, skipped_missing = 0, 0
            with transaction.atomic():
                for r in records:
                    sp = r.student
                    if not sp:
                        skipped_missing += 1
                        continue
                    sp.classroom = r.classroom
                    sp.section = r.section
                    sp.save(update_fields=['classroom', 'section'])
                    reverted += 1
            return Response({
                "status": "ok",
                "school": school.id,
                "year": str(year),
                "reverted": reverted,
                "skipped_missing": skipped_missing,
                "target_class_id": int(target_class_id) if target_class_id else None
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"detail": f"Failed to revert promotion: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def year_report(self, request):
        school_id = request.query_params.get('school')
        year = request.query_params.get('year')
        classroom_id = request.query_params.get('classroom')
        section_id = request.query_params.get('section')
        if not school_id or not year:
            return Response({"detail": "school and year are required"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            from academics.models import StudentYearRecord, StudentProfile
            qs = StudentYearRecord.objects.select_related('student', 'classroom', 'section').filter(school_id=school_id, academic_year=str(year))
            if classroom_id:
                qs = qs.filter(classroom_id=classroom_id)
            if section_id:
                qs = qs.filter(section_id=section_id)
            total = qs.count()
            promoted = qs.filter(status='promoted').count()
            retained = qs.filter(status='retained').count()
            not_passed = qs.filter(status='not_passed').count()
            cgpas = list(qs.exclude(result_cgpa__isnull=True).values_list('result_cgpa', flat=True))
            percentages = list(qs.exclude(percentage__isnull=True).values_list('percentage', flat=True))
            avg_cgpa = (sum(float(x) for x in cgpas) / len(cgpas)) if cgpas else None
            avg_percentage = (sum(float(x) for x in percentages) / len(percentages)) if percentages else None
            buckets = {'0-1': 0, '1-2': 0, '2-3': 0, '3-4': 0, '4-5': 0}
            for x in cgpas:
                v = float(x)
                if v < 1: buckets['0-1'] += 1
                elif v < 2: buckets['1-2'] += 1
                elif v < 3: buckets['2-3'] += 1
                elif v < 4: buckets['3-4'] += 1
                else: buckets['4-5'] += 1
            records = []
            for r in qs:
                s = r.student
                records.append({
                    'id': r.id,
                    'student_id': s.id,
                    'student_name': f"{getattr(s.user, 'first_name', '')} {getattr(s.user, 'last_name', '')}".strip() or getattr(s.user, 'username', ''),
                    'classroom': getattr(r.classroom, 'name', None),
                    'section': getattr(r.section, 'name', None),
                    'roll_number': r.roll_number,
                    'status': r.status,
                    'result_cgpa': r.result_cgpa,
                    'result_grade': r.result_grade,
                    'percentage': r.percentage,
                    'rank': r.rank,
                    'promoted_to_classroom': getattr(r.promoted_to_classroom, 'name', None)
                })
            data = {
                'school': int(school_id),
                'year': str(year),
                'classroom': int(classroom_id) if classroom_id else None,
                'section': int(section_id) if section_id else None,
                'summary': {
                    'total': total,
                    'promoted': promoted,
                    'retained': retained,
                    'not_passed': not_passed,
                    'avg_cgpa': avg_cgpa,
                    'avg_percentage': avg_percentage,
                    'cgpa_buckets': [{'label': k, 'count': v} for k, v in buckets.items()]
                },
                'records': records
            }
            return Response(data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"detail": "year report failed"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class TeacherAssignmentViewSet(viewsets.ModelViewSet):
    queryset = TeacherAssignment.objects.select_related('teacher','subject','classroom','section').all()
    serializer_class = TeacherAssignmentSerializer
    permission_classes = [AdminOrReadOnly]
    filter_backends = [filters.SearchFilter]
    search_fields = ['teacher__username','teacher__first_name','teacher__last_name']
    
    def get_queryset(self):
        queryset = TeacherAssignment.objects.select_related('teacher','subject','classroom','section').all()
        classroom_id = self.request.query_params.get('classroom')
        school_id = (self.request.query_params.get('classroom__school') 
                     or self.request.query_params.get('school') 
                     or self.request.query_params.get('school_id'))
        teacher_id = self.request.query_params.get('teacher') or self.request.query_params.get('teacher_id')
        subject_id = self.request.query_params.get('subject') or self.request.query_params.get('subject_id')
        if classroom_id:
            queryset = queryset.filter(classroom_id=classroom_id)
        if school_id:
            # Primary: filter by classroom's school
            queryset = queryset.filter(classroom__school_id=school_id)
            # Fallback: also restrict by subject's school to be safe
            queryset = queryset.filter(subject__school_id=school_id)
        if teacher_id:
            queryset = queryset.filter(teacher_id=teacher_id)
        if subject_id:
            queryset = queryset.filter(subject_id=subject_id)
        return queryset


# ---- Import Students API ----
from django.contrib.auth import get_user_model
from django.db import transaction
import io
import csv

User = get_user_model()


class ImportStudentsAPI(APIView):
    permission_classes = [AdminOrReadOnly]

    REQUIRED_COLUMNS = {"username", "first_name", "last_name", "classroom", "section", "roll_number"}

    def post(self, request):
        school_id = request.POST.get("school") or request.query_params.get("school")
        if not school_id:
            return Response({"detail": "Parameter 'school' is required."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            school = School.objects.get(pk=school_id)
        except School.DoesNotExist:
            return Response({"detail": "Invalid school id."}, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "No file uploaded. Use form field 'file'."}, status=status.HTTP_400_BAD_REQUEST)

        name = file.name.lower()
        try:
            if name.endswith(".csv"):
                created, updated, errors = self._import_csv(file, school)
            elif name.endswith(".docx"):
                created, updated, errors = self._import_docx(file, school)
            elif name.endswith(".pdf"):
                created, updated, errors = self._import_pdf(file, school)
            elif name.endswith((".xlsx", ".xlsm")):
                created, updated, errors = self._import_xlsx(file, school)
            elif name.endswith((".png", ".jpg", ".jpeg")):
                created, updated, errors = self._import_image(file, school)
            else:
                return Response({"detail": "Unsupported file type. Use CSV, DOCX, PDF, or image."}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"detail": f"Failed to import: {e}"}, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            "message": "Import complete",
            "created": created,
            "updated": updated,
            "errors": errors,
        }, status=status.HTTP_200_OK)

    def _normalize_headers(self, headers):
        return [h.strip().lower().replace(" ", "_") for h in headers]

    @transaction.atomic
    def _import_csv(self, uploaded_file, school):
        created, updated = 0, 0
        errors = []
        # Attempt utf-8-sig then fallback latin-1
        content = uploaded_file.read()
        for enc in ["utf-8-sig", "utf-8", "latin-1"]:
            try:
                text = content.decode(enc)
                break
            except Exception:
                continue
        else:
            raise ValueError("Unable to decode CSV file")

        reader = csv.DictReader(io.StringIO(text))
        reader.fieldnames = self._normalize_headers(reader.fieldnames or [])

        # Optional: allow partial columns
        supported = {"username", "first_name", "last_name", "password", "classroom", "section", "roll_number",
                     "parent", "guardian", "guardian_name", "father_name", "mother_name"}

        row_num = 1
        for row in reader:
            row_num += 1
            data = {k: (row.get(k) or "").strip() for k in supported}
            try:
                cu, uu = self._create_or_update_student(data, school)
                created += cu
                updated += uu
            except Exception as e:
                errors.append({"row": row_num, "error": str(e)})

        return created, updated, errors

    def _import_docx(self, uploaded_file, school):
        try:
            import docx  # python-docx
        except Exception:
            raise ValueError("python-docx not installed on server")
        # Very simple extraction: read table rows into dicts with header
        document = docx.Document(uploaded_file)
        created, updated, errors = 0, 0, []
        for table in document.tables:
            headers = self._normalize_headers([cell.text for cell in table.rows[0].cells]) if table.rows else []
            supported = {"username", "first_name", "last_name", "password", "classroom", "section", "roll_number",
                         "parent", "guardian", "guardian_name", "father_name", "mother_name"}
            for r_i, row in enumerate(table.rows[1:], start=2):
                values = [cell.text.strip() for cell in row.cells]
                data = {h: (values[i] if i < len(values) else "") for i, h in enumerate(headers) if h in supported}
                try:
                    cu, uu = self._create_or_update_student(data, school)
                    created += cu
                    updated += uu
                except Exception as e:
                    errors.append({"row": r_i, "error": str(e)})
        return created, updated, errors

    def _import_pdf(self, uploaded_file, school):
        # Try pdfplumber for table extraction
        try:
            import pdfplumber
        except Exception:
            raise ValueError("pdfplumber not installed on server")
        created, updated, errors = 0, 0, []
        with pdfplumber.open(uploaded_file) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables() or []
                for t in tables:
                    if not t or not t[0]:
                        continue
                    headers = self._normalize_headers(t[0])
                    supported = {"username", "first_name", "last_name", "password", "classroom", "section", "roll_number",
                                 "parent", "guardian", "guardian_name", "father_name", "mother_name"}
                    for r_i, row in enumerate(t[1:], start=2):
                        data = {h: (row[i].strip() if i < len(row) and row[i] else "") for i, h in enumerate(headers) if h in supported}
                        try:
                            cu, uu = self._create_or_update_student(data, school)
                            created += cu
                            updated += uu
                        except Exception as e:
                            errors.append({"row": r_i, "error": str(e)})
        return created, updated, errors

    def _import_xlsx(self, uploaded_file, school):
        try:
            from openpyxl import load_workbook
        except Exception:
            raise ValueError("openpyxl not installed on server")

        created, updated, errors = 0, 0, []
        try:
            wb = load_workbook(uploaded_file, data_only=True)
        except Exception as e:
            raise ValueError(f"Failed to open Excel file: {e}")

        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            return 0, 0, [{"row": 0, "error": "Empty Excel sheet"}]
        headers = self._normalize_headers([str(h) if h is not None else '' for h in headers])
        supported = {"username", "first_name", "last_name", "password", "classroom", "section", "roll_number",
                    "parent", "guardian", "guardian_name", "father_name", "mother_name"}

        for idx, row in enumerate(rows_iter, start=2):
            values = [str(v).strip() if v is not None else '' for v in row]
            data = {h: (values[i] if i < len(values) else '') for i, h in enumerate(headers) if h in supported}
            try:
                cu, uu = self._create_or_update_student(data, school)
                created += cu
                updated += uu
            except Exception as e:
                errors.append({"row": idx, "error": str(e)})
        return created, updated, errors

    def _import_image(self, uploaded_file, school):
        """
        Detect table grid and OCR per-cell to map 6 columns:
        [serial, parent, student, class, section, roll_number]
        Uses OpenCV (cv2) + Tesseract (ben+eng). Falls back to naive OCR if grid detection fails.
        """
        try:
            from PIL import Image
            import pytesseract
            # Set Tesseract path for Windows
            import os
            if os.name == 'nt':  # Windows
                tesseract_path = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
                if os.path.exists(tesseract_path):
                    pytesseract.pytesseract.tesseract_cmd = tesseract_path
        except Exception:
            raise ValueError("pillow/pytesseract not installed on server")

        # Try advanced grid-based extraction
        try:
            import cv2
            import numpy as np
        except Exception:
            cv2 = None
            np = None

        created, updated, errors = 0, 0, []

        # Load image
        image = Image.open(uploaded_file).convert('RGB')

        def ocr_cell(img_pil):
            try:
                # Bengali + English; page segmentation mode 7 (single text line) works well for cells
                return pytesseract.image_to_string(
                    img_pil,
                    lang='ben+eng',
                    config='--psm 7'
                ).strip()
            except Exception:
                return ''

        used_advanced = False
        if cv2 is not None and np is not None:
            try:
                # Convert to OpenCV format
                img_cv = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
                gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
                # Adaptive threshold for robust binarization
                thr = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_MEAN_C,
                                            cv2.THRESH_BINARY_INV, 15, 10)

                # Detect horizontal and vertical lines
                h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (40, 1))
                v_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 40))
                h_lines = cv2.morphologyEx(thr, cv2.MORPH_OPEN, h_kernel, iterations=2)
                v_lines = cv2.morphologyEx(thr, cv2.MORPH_OPEN, v_kernel, iterations=2)
                table_mask = cv2.add(h_lines, v_lines)

                # Find contours of boxes (cells)
                contours, _ = cv2.findContours(table_mask, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
                boxes = []
                for cnt in contours:
                    x, y, w, h = cv2.boundingRect(cnt)
                    # Filter too small or too large boxes
                    if w < 40 or h < 20:
                        continue
                    boxes.append((y, x, w, h))
                if not boxes:
                    raise RuntimeError('No boxes detected')

                # Sort boxes top-to-bottom, then left-to-right
                boxes.sort()
                # Group boxes into rows by y proximity
                rows = []
                row = []
                last_y = None
                tol = 12  # vertical tolerance
                for (y, x, w, h) in boxes:
                    if last_y is None or abs(y - last_y) <= tol:
                        row.append((y, x, w, h))
                        last_y = y
                    else:
                        rows.append(sorted(row, key=lambda b: b[1]))
                        row = [(y, x, w, h)]
                        last_y = y
                if row:
                    rows.append(sorted(row, key=lambda b: b[1]))

                # Heuristic: skip the top header row(s) by requiring first cell to be a number
                processed_rows = 0
                for r_idx, r in enumerate(rows, start=1):
                    # Expect at least 6 columns; if more, we take first 6
                    if len(r) < 6:
                        continue
                    cells = r[:6]
                    texts = []
                    for (y, x, w, h) in cells:
                        crop = img_cv[y:y+h, x:x+w]
                        crop = cv2.cvtColor(crop, cv2.COLOR_BGR2RGB)
                        crop_pil = Image.fromarray(crop)
                        texts.append(ocr_cell(crop_pil))

                    # Validate first cell numeric
                    serial = ''.join(ch for ch in texts[0] if ch.isdigit())
                    if not serial:
                        # likely header
                        continue

                    parent = texts[1]
                    student = texts[2]
                    classroom = texts[3]
                    section = texts[4]
                    roll_number = ''.join(ch for ch in texts[5] if ch.isdigit()) or texts[5]

                    # Map to creation data
                    data = {
                        # We'll set username empty to auto-generate from student
                        "username": "",
                        "first_name": student.split()[0] if student else "",
                        "last_name": ' '.join(student.split()[1:]) if student and len(student.split()) > 1 else "",
                        "password": "",
                        "classroom": classroom,
                        "section": section,
                        "roll_number": roll_number,
                        "guardian_name": parent,
                    }
                    try:
                        cu, uu = self._create_or_update_student(data, school)
                        created += cu
                        updated += uu
                        processed_rows += 1
                    except Exception as e:
                        errors.append({"row": r_idx, "error": str(e)})

                if processed_rows == 0:
                    raise RuntimeError('No data rows recognized from table')

                used_advanced = True
            except Exception as e:
                # Fall back to naive OCR parsing
                used_advanced = False

        if not used_advanced:
            # Fallback: naive whole-image OCR and parse by multiple spaces or commas
            try:
                text = pytesseract.image_to_string(image, lang='ben+eng', config='--psm 6')
            except Exception as e:
                raise ValueError(f"OCR failed: {e}")

            lines = [l for l in (text.splitlines()) if l and len(l.strip()) > 0]
            # Try to find the first line with a leading number to start data rows
            started = False
            row_idx = 0
            for l in lines:
                parts = [p for p in l.strip().split('\t') if p]  # sometimes tesseract uses tabs
                if len(parts) < 2:
                    # split by 2+ spaces
                    parts = [p for p in filter(None, [p.strip() for p in __import__('re').split(r"\s{2,}", l)])]
                # Expect at least 6 columns; if more, keep first 6
                if len(parts) >= 2 and (parts[0].strip().isdigit() or parts[0].strip().replace('.', '').isdigit()):
                    started = True
                if not started:
                    continue
                row_idx += 1
                if len(parts) < 6:
                    # can't parse; keep as error
                    errors.append({"row": row_idx, "error": f"Unparsable row: {l}"})
                    continue
                cols = parts[:6]
                serial = cols[0]
                parent = cols[1]
                student = cols[2]
                classroom = cols[3]
                section = cols[4]
                roll_number = ''.join(ch for ch in cols[5] if ch.isdigit()) or cols[5]

                data = {
                    "username": "",
                    "first_name": student.split()[0] if student else "",
                    "last_name": ' '.join(student.split()[1:]) if student and len(student.split()) > 1 else "",
                    "password": "",
                    "classroom": classroom,
                    "section": section,
                    "roll_number": roll_number,
                    "guardian_name": parent,
                }
                try:
                    cu, uu = self._create_or_update_student(data, school)
                    created += cu
                    updated += uu
                except Exception as e:
                    errors.append({"row": row_idx, "error": str(e)})

        return created, updated, errors

    def _create_or_update_student(self, data, school):
        """
        Create or update a student and StudentProfile. Returns (created_count, updated_count)
        Expected fields in data: username, first_name, last_name, password, classroom, section, roll_number,
        and optional guardian fields (guardian_name/parent/father_name/mother_name).
        Also auto-creates a Parent user+Profile when guardian_name is provided.
        """
        username = data.get("username") or ""
        first_name = data.get("first_name") or ""
        last_name = data.get("last_name") or ""
        password = data.get("password") or ""
        classroom_name = data.get("classroom") or ""
        section_name = data.get("section") or ""
        roll_number = data.get("roll_number") or ""
        guardian_name = (
            data.get("guardian_name") or data.get("guardian") or data.get("parent")
            or data.get("father_name") or data.get("mother_name") or ""
        ).strip()

        if not username and not first_name:
            raise ValueError("Either username or first_name is required")

        # Ensure username
        if not username:
            base = (first_name or "student").lower().replace(" ", "")
            candidate = base
            idx = 1
            while User.objects.filter(username=candidate).exists():
                idx += 1
                candidate = f"{base}{idx}"
            username = candidate

        user, created_user = User.objects.get_or_create(username=username, defaults={
            "first_name": first_name,
            "last_name": last_name,
        })
        # If user existed, optionally update names
        if not created_user:
            changed = False
            if first_name and user.first_name != first_name:
                user.first_name = first_name; changed = True
            if last_name and user.last_name != last_name:
                user.last_name = last_name; changed = True
            if changed:
                user.save()
        if password:
            user.set_password(password)
            user.save()

        # Classroom and Section
        classroom = None
        if classroom_name:
            classroom, _ = ClassRoom.objects.get_or_create(school=school, name=classroom_name)
        section = None
        if classroom and section_name:
            section, _ = Section.objects.get_or_create(classroom=classroom, name=section_name)

        # Prepare guardian user if provided
        guardian_user = None
        if guardian_name:
            # Create or reuse a guardian user by generated username
            base = guardian_name.replace(' ', '').lower() or 'parent'
            candidate = base
            idx = 1
            from django.contrib.auth import get_user_model
            U = get_user_model()
            while U.objects.filter(username=candidate).exists():
                idx += 1
                candidate = f"{base}{idx}"
            guardian_user, created_g = U.objects.get_or_create(username=candidate, defaults={
                "first_name": guardian_name.split()[0] if guardian_name else "",
                "last_name": ' '.join(guardian_name.split()[1:]) if guardian_name and len(guardian_name.split()) > 1 else "",
            })
            # Ensure Profile with role parent
            from users.models import Profile as UserProfile
            UserProfile.objects.update_or_create(user=guardian_user, defaults={"school": school, "role": "parent"})

        sp, created_profile = StudentProfile.objects.get_or_create(user=user, defaults={
            "school": school,
            "classroom": classroom,
            "section": section,
            "roll_number": roll_number or None,
            "guardian_name": guardian_name or None,
            "guardian": guardian_user,
        })
        if not created_profile:
            changed = False
            if sp.school_id != school.id:
                sp.school = school; changed = True
            if classroom and sp.classroom_id != classroom.id:
                sp.classroom = classroom; changed = True
            if section and (sp.section_id or None) != (section.id if section else None):
                sp.section = section; changed = True
            if roll_number and sp.roll_number != roll_number:
                sp.roll_number = roll_number; changed = True
            if guardian_name and sp.guardian_name != guardian_name:
                sp.guardian_name = guardian_name; changed = True
            if guardian_user and (sp.guardian_id or None) != guardian_user.id:
                sp.guardian = guardian_user; changed = True
            if changed:
                sp.save()
            return 0, 1

        return 1, 0
